from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font
import os
from datetime import datetime, timedelta

from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

from constants import row_heights, col_widths

class Label:
    ROWS_PER_LABEL = 17

    def __init__(self, ws, start_row, label_data):
        self.ws = ws
        self.start_row = start_row
        self.row_offset = start_row - 1
        self.label_data = label_data

        self.merge_ranges = [
            "A1:E8", "A9:B12", "C9:E12", "A13:E16", "F1:L4", "M1:O4", "P1:R4", "S1:S16",
            "F5:I8", "J5:L8", "M5:O8", "P5:R8", "F9:I12", "J9:L12", "M9:O12", "P9:R12",
            "F13:G14", "H13:I14", "J13:K14", "F15:G16", "H15:I16", "J15:K16", "L13:M16",
            "N13:N16", "O13:O16", "P13:R16"
        ]

        self.images_info = [
            ("images/Logo.png", "A2", 323.62, 108.4615384615385),
            ("images/EAC.png", "A9", 77.214, 61.53856),
            ("images/Contacts.png", "C9", 193.035, 65.38455),
        ]

        self.center_alignment = Alignment(horizontal="center", vertical="center")

        self.thick_border = Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        )

    def _apply_row_heights(self):
        for r, h in row_heights.items():
            self.ws.row_dimensions[r + self.row_offset].height = h

    def _apply_merge_and_borders(self):
        for merge_range in self.merge_ranges:
            start_cell, end_cell = merge_range.split(':')
            start_col_letter, start_row = coordinate_from_string(start_cell)
            end_col_letter, end_row = coordinate_from_string(end_cell)
            start_col = column_index_from_string(start_col_letter)
            end_col = column_index_from_string(end_col_letter)

            new_start_cell = f"{start_col_letter}{start_row + self.row_offset}"
            new_end_cell = f"{end_col_letter}{end_row + self.row_offset}"
            new_merge_range = f"{new_start_cell}:{new_end_cell}"

            self.ws.merge_cells(new_merge_range)

            for row in range(start_row + self.row_offset, end_row + self.row_offset + 1):
                for col in range(start_col, end_col + 1):
                    cell = self.ws.cell(row=row, column=col)
                    cell.border = self.thick_border

    def _insert_images(self):
        for path, cell, width, height in self.images_info:
            if not os.path.exists(path):
                continue

            try:
                col_letter, row_num = coordinate_from_string(cell)
                new_row = row_num + self.row_offset
                new_cell = f"{col_letter}{new_row}"
                img = Image(path)
                img.width = width
                img.height = height
                self.ws.add_image(img, new_cell)
            except Exception as e:
                print(f"Ошибка при вставке изображения {path}: {e}")

    def _set_text_cells(self):
        data = self.label_data
        dimensions = data.get('dimensions', (0, 0, 0))

        # Определяем значение для J9
        label_type = data['label_type'].upper()
        if label_type == "КОРПУС":
            j9_value = data['carcase']
        elif label_type == "ОРГАЛИТ":
            j9_value = "БЕЛЫЙ"
        elif label_type in ["ФАСАДЫ МДФ", "ФАСАДЫ ПЛАСТИК"]:
            j9_value = data.get('facade', '')
        else:
            j9_value = data.get('extra_component', '')

        # Формируем текст для M9
        m9_text = f"{data.get('client', '')}/{data.get('store_number', '')}"

        text_cells = [
            ("A13", "ГОСТ 16371-2014", 16),
            ("F1", data.get('item_name', ''), 16),
            ("F9", data['label_type'].upper(), 24),
            ("J9", j9_value, 16),
            ("F15", str(dimensions[1]) if len(dimensions) > 1 else "", 14),
            ("H15", str(dimensions[0]) if len(dimensions) > 0 else "", 14),
            ("J15", str(dimensions[2]) if len(dimensions) > 2 else "", 14),
            ("N13", str(int(data['weight'])) if data.get('weight') else "", 14),
            ("M1", f"№ {data.get('order_number', '')}", 20),
            ("M9", m9_text, 14),
            ("P5", str(data.get('package_total', 1)), 20),
            ("P13", str(data.get('package_num', 1)), 20),
            ("F5", "Наименование упаковки", 16),
            ("J5", "Цвет", 20),
            ("M5", "ЗАКАЗЧИК", 20),
            ("P1", "ВСЕГО УПАКОВОК", 14),
            ("P9", "№ УПАКОВКИ", 14),
            ("F13", "ВЫСОТА", 14),
            ("H13", "ШИРИНА", 14),
            ("J13", "ГЛУБИНА", 14),
            ("L13", "ВЕС", 14),
            ("O13", "КГ", 14),
        ]

        for cell, text, size in text_cells:
            if not text:
                continue

            col_letter, row_num = coordinate_from_string(cell)
            new_row = row_num + self.row_offset
            new_cell = f"{col_letter}{new_row}"

            self.ws[new_cell] = text
            self.ws[new_cell].font = Font(name="Times New Roman", size=size, bold=True)
            self.ws[new_cell].alignment = self.center_alignment

    def _set_date(self):
        date_cell = f"S{1 + self.row_offset}"
        label_date = (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y")
        self.ws[date_cell] = label_date
        self.ws[date_cell].font = Font(name="Times New Roman", size=26, bold=True)
        self.ws[date_cell].alignment = Alignment(
            horizontal="center", vertical="center", textRotation=90
        )

    def create(self):
        self._apply_row_heights()
        self._apply_merge_and_borders()
        self._insert_images()
        self._set_text_cells()
        self._set_date()
